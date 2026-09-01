"""Excel 매출 실행계획 파일 파싱 및 정규화.

Streamlit 의존 없음. `load_plan(bytes)` 하나가 공개 진입점이다.
"""

from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from . import schema

#: Excel 날짜 serial 의 기준일. 1900 윤년 버그 때문에 1899-12-31 이 아니라 12-30 이다.
_EXCEL_EPOCH = _dt.datetime(1899, 12, 30)

#: 차원 값이 이 중 하나면 "미지정" 으로 통일한다.
#: 원본에서 빈 셀이 숫자 0 으로 저장된 경우가 있어 0 도 포함한다(예: 비고 열).
_BLANK_TOKENS = {"", "0", "0.0", "none", "nan", "-"}


class PlanLoadError(Exception):
    """파일을 읽을 수 없거나 양식을 인식할 수 없을 때."""


@dataclass
class LoadResult:
    """파싱 결과와 그 과정에서 알아낸 메타정보."""

    df: pd.DataFrame
    sheet_name: str
    header_row: int
    row_count: int
    period_min: str
    period_max: str
    periods: list[str]
    source_name: str = ""
    warnings: list[str] = field(default_factory=list)

    @property
    def period_range(self) -> str:
        if not self.periods:
            return "-"
        if self.period_min == self.period_max:
            return self.period_min
        return f"{self.period_min} ~ {self.period_max}"


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def load_plan(data: bytes, source_name: str = "") -> LoadResult:
    """xlsx 바이트를 정규화된 DataFrame 으로 변환한다."""
    warnings: list[str] = []

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # 손상/암호화 파일
        raise PlanLoadError(f"Excel 파일을 열 수 없습니다: {exc}") from exc

    try:
        ws, sheet_name = _pick_sheet(wb, warnings)
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not rows:
        raise PlanLoadError("시트가 비어 있습니다.")

    header_row = _find_header_row(rows)
    header_by_letter = _header_by_letter(rows[header_row - 1])
    warnings.extend(schema.validate_headers(header_by_letter))

    df = _build_frame(rows, header_row)
    if df.empty:
        raise PlanLoadError("헤더 아래에 데이터 행이 없습니다.")

    df = _normalize(df, warnings)

    periods = sorted(df[schema.PERIOD_COLUMN].unique().tolist())
    return LoadResult(
        df=df,
        sheet_name=sheet_name,
        header_row=header_row,
        row_count=len(df),
        period_min=periods[0] if periods else "-",
        period_max=periods[-1] if periods else "-",
        periods=periods,
        source_name=source_name,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# 내부 단계
# ---------------------------------------------------------------------------

def _pick_sheet(wb, warnings: list[str]):
    if schema.SHEET_NAME in wb.sheetnames:
        return wb[schema.SHEET_NAME], schema.SHEET_NAME
    fallback = wb.sheetnames[0]
    warnings.append(
        f"'{schema.SHEET_NAME}' 시트가 없어 첫 번째 시트 '{fallback}' 를 사용합니다."
    )
    return wb[fallback], fallback


def _find_header_row(rows: list[tuple]) -> int:
    """A열 값이 '변경'인 행을 헤더로 본다(1-based). 행 번호 하드코딩 금지."""
    marker_idx = column_index_from_string(schema.HEADER_MARKER_COLUMN) - 1
    limit = min(len(rows), schema.HEADER_SEARCH_LIMIT)
    for i in range(limit):
        cell = rows[i][marker_idx] if marker_idx < len(rows[i]) else None
        if cell is not None and str(cell).strip() == schema.HEADER_MARKER_VALUE:
            return i + 1
    raise PlanLoadError(
        f"헤더 행을 찾지 못했습니다. 상위 {limit}행의 "
        f"{schema.HEADER_MARKER_COLUMN}열에 '{schema.HEADER_MARKER_VALUE}' 가 없습니다."
    )


def _header_by_letter(header_row_values: tuple) -> dict[str, object]:
    out: dict[str, object] = {}
    for letter in schema.EXPECTED_HEADERS:
        idx = column_index_from_string(letter) - 1
        out[letter] = header_row_values[idx] if idx < len(header_row_values) else None
    return out


def _build_frame(rows: list[tuple], header_row: int) -> pd.DataFrame:
    """열 문자 기준으로 표준 컬럼명 DataFrame 을 만든다."""
    letters = list(schema.COLUMN_MAP)
    indices = [column_index_from_string(x) - 1 for x in letters]
    names = [schema.COLUMN_MAP[x] for x in letters]
    width = max(indices) + 1

    records = []
    for row in rows[header_row:]:
        if _is_blank_row(row):
            continue
        padded = row + (None,) * (width - len(row)) if len(row) < width else row
        records.append([padded[i] for i in indices])

    return pd.DataFrame(records, columns=names)


def _is_blank_row(row: tuple) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


def _normalize(df: pd.DataFrame, warnings: list[str]) -> pd.DataFrame:
    df = df.copy()

    for col in schema.DIMENSIONS:
        df[col] = df[col].map(_clean_dimension)

    df[schema.PERIOD_COLUMN] = df[schema.PERIOD_COLUMN].map(_to_period)
    bad = df[schema.PERIOD_COLUMN].isna()
    if bad.any():
        warnings.append(f"기준년월을 해석할 수 없는 행 {int(bad.sum())}건을 제외했습니다.")
        df = df[~bad].copy()

    for col in schema.NUMERIC_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    period_ts = pd.to_datetime(df[schema.PERIOD_COLUMN], format="%Y-%m")
    df["연도"] = period_ts.dt.year
    df["월"] = period_ts.dt.month

    df[schema.QUARTER_COLUMN] = _coerce_period_part(
        df[schema.QUARTER_COLUMN], derived=(df["월"] - 1) // 3 + 1, valid={1, 2, 3, 4},
        label="분기", warnings=warnings,
    )
    df[schema.HALF_COLUMN] = _coerce_period_part(
        df[schema.HALF_COLUMN], derived=(df["월"] - 1) // 6 + 1, valid={1, 2},
        label="반기", warnings=warnings,
    )

    return df.reset_index(drop=True)


def _coerce_period_part(original, derived, valid: set[int], label: str,
                        warnings: list[str]):
    """원본 분기/반기 값을 우선 쓰되(FR-302), 기준년월과 대조해 검증한다.

    원본 값이 비었거나, 범위를 벗어났거나, **기준년월에서 파생한 값과 다르면**
    파생값으로 대체하고 경고를 남긴다.

    이 검증이 필요한 이유: 실제 샘플 데이터에 4월인데 분기=1, 7월인데 분기=1 로
    찍힌 행이 섞여 있다. 이를 그대로 믿으면 '월 합계 == 분기 합계' 가 깨져
    집계 결과를 신뢰할 수 없게 된다.
    """
    derived = derived.astype("int64")
    numeric = pd.to_numeric(original, errors="coerce")

    in_range = numeric.isin(list(valid))
    consistent = in_range & (numeric == derived)

    n_missing = int((~in_range).sum())
    if n_missing:
        warnings.append(
            f"{label} 값이 비었거나 범위를 벗어난 행 {n_missing}건은 기준년월에서 파생했습니다."
        )

    conflicted = in_range & ~consistent
    n_conflict = int(conflicted.sum())
    if n_conflict:
        warnings.append(
            f"{label} 값이 기준년월과 불일치하는 행 {n_conflict}건을 기준년월 기준으로 보정했습니다."
        )

    return numeric.where(consistent, derived).astype("int64")


def _clean_dimension(value) -> str:
    """차원 값을 문자열로 통일. 빈값·0 은 '미지정'."""
    if value is None:
        return "미지정"
    if isinstance(value, float) and pd.isna(value):
        return "미지정"
    if isinstance(value, (int, float)) and float(value) == 0.0:
        return "미지정"
    text = " ".join(str(value).split())
    if text.lower() in _BLANK_TOKENS:
        return "미지정"
    return text


def _to_period(value) -> str | None:
    """기준년월을 'YYYY-MM' 문자열로. Excel serial / datetime / 문자열 모두 처리."""
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, _dt.date):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, (int, float)):
        if pd.isna(value) or value <= 0:
            return None
        dt = _EXCEL_EPOCH + _dt.timedelta(days=float(value))
        return f"{dt.year:04d}-{dt.month:02d}"
    text = str(value).strip()
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return f"{parsed.year:04d}-{parsed.month:02d}"


